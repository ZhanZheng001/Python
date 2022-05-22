import openpyxl
from openpyxl import Workbook

wb=openpyxl.load_workbook('test.xlsx')
index = wb.sheetnames
ws = wb[index[2]]
# print(ws['A1'].value)
#
for row in ws.rows:
    data = [cell.value for cell in row]
    print(data)


print(ws.max_row)
print(ws.max_column)

#迭代行
# for row  in ws.iter_rows(min_row=2,min_col=2,max_row=ws.max_row,max_col=ws.max_column):
#     data=[cell.value for cell in row]
#     print(data)
#列迭代
for column in ws.iter_cols(min_row=2,min_col=1,max_row=ws.max_row,max_col=ws.max_column):
    data = [cell.value for cell in column]
    print(data)

# wb = Workbook()
# ws = wb.active
# ws1 = wb.create_sheet('test.xlsx',0)
# ws2 = wb.create_sheet('用户信息.xlsx',2)
#
# ws1['A1'] = 'aaa'
# ws1['B1'] = 'bbb'
# ws1['A2'] = '222'
# ws1.cell(row = 5,column = 1).value='大熊课堂'
#
# #追加方式
# rows = (
#     ('andy','男',18),
#     ('Jack','男',50),
#     ('张三','男',20)
# )
# for row in rows:
#     ws2.append(row)
#
# wb.save('test.xlsx')