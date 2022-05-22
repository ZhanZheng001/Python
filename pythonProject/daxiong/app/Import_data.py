import os
from flask_script import Command
import openpyxl
from app.config import BASE_DIR
from app.models import Course,Sale
from app import db

class ImportData(Command):
    def run(self):
        print('导入开始')
        dir = os.path.join(BASE_DIR,'c_excel')
        for file_name in os.listdir(dir):
            file_path = os.path.join(dir,file_name)
            self.save_to_mysql(file_path)
        print('导入结束')

    def save_to_mysql(self,file_path):
        fileds = ['productId','courseId','productName','productType','provider',
             'score','scoreLevel','learnerCount','lessonCount','couponPrice',
             'originalPrice','discountPrice','vipPrice','imgUrl','bigImgUrl',
             'description']
        create_time = file_path.split('.')[0][-10:]
        #导入excel
        wb = openpyxl.load_workbook(file_path)
        #选择worksheet
        index = wb.sheetnames
        ws = wb[index[0]]
        #迭代行
        for row in ws.iter_rows(min_row=2):
            data=[cell.value for cell in row]
            dict_val=dict(zip(fileds,data))
            #存入course表
            course = Course(**dict_val)
            self.save_course(course)
            #存入sale表
            sale = Sale(productId = dict_val['productId'],productName=dict_val['productName'],
                        learnerCount=dict_val['learnerCount'],create_time=create_time)
            self.save_sale(sale,create_time)
        wb.close()

    def save_course(self,course):
        try:
            db.session.merge(course)  # merge包括新增和更改
            db.session.commit()
        except:
            db.session.rollback()
    def save_sale(self,sale,create_time):
        try:
            data = Sale.query.filter_by(productId=sale.productId, create_time=create_time).first()
            if not data:
                db.session.add(sale)
                db.session.commit()
        except:
            db.session.rollback()

if __name__=='__main__':
    print(BASE_DIR)